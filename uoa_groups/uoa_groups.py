"""

A library to help with University of Auckland group management.

Requirements:

 - argparse
 - openpyxl

Install from git:

    TODO

"""


import os.path
import logging
import uoa_ldap

from openpyxl import Workbook, load_workbook
from openpyxl.cell import get_column_letter, column_index_from_string

def filter_duplicate_groups(list_of_groups):
    '''
    Filters out duplicate groups (if already contained in a tree-branch that goes higher up).
    '''
    filtered = list(list_of_groups)
    for g in list_of_groups:
        for h in list_of_groups:
            if h.is_child_of(g):
                try:
                    filtered.remove(g)
                except ValueError:
                    pass
                break
            
    return filtered

class UoA_group(object):

    '''
    Class to describe a group within the UoA group hierarchy.
    '''

    def __init__(self, gid, name):
        self.gid = gid
        self.name = name
        self.parent = None
        self.childs = []

    def __str__(self):
        return self.gid+' ('+self.name+')'

    def __repr__(self):
        return self.__str__()

    def __eq__(self, other):
        # assuming that there is no typo in the excel document where
        # abbreviation and pretty name differ for the same group
        return self.gid == other.gid

    def add_child(self, gid, name):
        """Adds a child with the specified group id and name."""

        child = UoA_group(gid, name)
        if not child in  self.childs:
            logging.info("Adding: "+str(gid))
            child.parent = self
            self.childs.append(child)

    def print_tree(self, ident=""):
        """Prints the hierarchy with this group as root."""
        print ident+str(self)
        for c in self.childs:
            c.print_tree(ident+"  ")
        
    def print_tree_down(self, ident=""):
        """Prints the parent hierarchy of this group"""

        if self.parent:
            ident = self.parent.print_tree_down(ident)

        print ident+str(self)
        return ident+"  "
        
            
    def get_child(self, gid, ignore_case=True):
        """Returns the child with the specified gid."""

        if not ignore_case:
            if self.gid == gid:
                return self
        else:
            if self.gid.lower() == gid.lower():
                return self

        for c in self.childs:
            found = c.get_child(gid, ignore_case)
            if found:
                return found

        return None

    
    def find_childs(self, search_term, search_matches, ignore_case=False):
        """Returns all childs that match the specified search term in either the id or name."""

        if ignore_case:
            search_term = search_term.lower()
            gid = self.gid.lower()
            name = self.name.lower()
        else:
            gid = self.gid
            name = self.name
        
        if search_term in gid or search_term in name:
            search_matches.append(self)

        for c in self.childs:
            search_matches = c.find_childs(search_term, search_matches, ignore_case)

        return search_matches


    def is_root_group(self):
        """Returns True if this group does not have a parent."""
        return self.parent == None

    def is_child_of(self, other):
        """Returns True if this group is a child of the provided group."""
        if not self.parent:
            return False

        if self.parent == other:
            return True

        return self.parent.is_child_of(other)
        
class UoA_groups(object):
    '''
    Base class to create and encapsulate the UoA hierarchy.
    '''
    
    def __init__(self, excel_file):
        '''Parsed using an Excel file provided by UoA HR.'''
        
        self.excel_file = excel_file
        self.wb = load_workbook(excel_file)
        self.sheet = self.wb.get_sheet_by_name("Data")

        self.root_gid = self.sheet['A2'].value
        self.root_name = "University of Auckland"
        self.root = UoA_group(self.root_gid, self.root_name)
        
        for row in range(2, self.sheet.max_row + 1):
            l1 = self.sheet['A'+str(row)].value
            if not l1 == self.root_gid:
                raise Exception("Error in spreadsheet: "+str(self.root_gid)+" != "+str(l1))
            l1_desc = self.root_name
            l1_group = UoA_group(l1, l1_desc)

            l2 = self.sheet['B'+str(row)].value
            if not l2:
                continue
            l2_desc = self.sheet['C'+str(row)].value
            
            self.root.add_child(l2, l2_desc)
            l2_group = self.root.get_child(l2)

            if not l2_group:
                self.root.print_tree()
                raise Exception("No group 2: "+str(l2))

            l3 = self.sheet['D'+str(row)].value
            if not l3:
                continue
            l3_desc = self.sheet['E'+str(row)].value

            l2_group.add_child(l3, l3_desc)
            l3_group = self.root.get_child(l3)

            if not l3_group:
                self.root.print_tree()
                raise Exception("No group 3: "+str(l3))

            
            l4 = self.sheet['F'+str(row)].value
            if not l4:
                continue
            l4_desc = self.sheet['G'+str(row)].value

            l3_group.add_child(l4, l4_desc)
            l4_group = self.root.get_child(l4)

            if not l4_group:
                self.root.print_tree()
                raise Exception("No group 4: "+str(l4))

            
            l5 = self.sheet['H'+str(row)].value
            if not l5:
                continue
            l5_desc = self.sheet['I'+str(row)].value

            l4_group.add_child(l5, l5_desc)
        
            
    def print_tree(self):
        """Prints the entire group hierarchy."""
        self.root.print_tree()

    def get_group(self, gid, ignore_case=False):
        """Finds the group with the specified group id or None if it doesn't exist."""

        return self.root.get_child(gid, ignore_case)

    def find_groups(self, search_term, ignore_case=False):
        """Finds all the groups that match the provided string in either the id or name."""

        matches = self.root.find_childs(search_term, [], ignore_case)

        # decided not to filter out groups in this case
        # return filter_duplicate_groups(matches)
        return matches

    def get_high_level_groups(self, list_of_group_ids):
        '''Filter out group tree-branches that are already part of one or more, higher-level group tree-branches.'''

        groups = [g for g in (self.get_group(gid) for gid in list_of_group_ids)]

        return filter_duplicate_groups(groups)


